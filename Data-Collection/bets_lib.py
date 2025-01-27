import datetime
import numpy as np
np.float = float    
np.int = int   #module 'numpy' has no attribute 'int'
np.object = object    #module 'numpy' has no attribute 'object'
np.bool = bool    
import pandas as pd 
import time
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options 
from selenium.webdriver.support.select import Select
from bs4 import BeautifulSoup
import re
from datetime import datetime,timedelta
from datetime import date
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os.path
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pyautogui as pag
import shutil
from os.path import exists




class Match():
    def __init__(self):
        self.is_singles = True
        self.partner = ''
        self.opponent = ''
        self.opponent_rank = ''
        self.scores = ''
        self.date = ''
        self.tournament = ''
        self.round = ''
        self.won = True
        self.in_three = False
        self.level = ''
        self.age_group = ''
        self.player_rank = ''
        self.partner_rank = ''
        self.is_level_dubs = ''
        self.woman_rank = ''
        self.man_rank = ''
        self.ranks_used = ''
        self.dicipline = ''

    def print_all_info(self):
        print(f'Tournament: {self.tournament}. Played on {self.date}.')
        print(f'They Played {self.opponent}, who were ranked {self.opponent_rank} at the time.')
        print(f'The scores were:{self.scores} and did they win? {self.won}... and in three? {self.in_three}')

class MatchClump():
    def __init__(self,games,week,link):
        self.games = games
        self.week_value = week
        self.link = link
#Helper Function: called ALOT
# BE Tournament software always needs you to accept cookies when opening for the first time in a new browser. 
# This function deals with the popup, should the popup pop up... Called often incase driver is closed and re-opened.
def deal_with_cookies(driver):
    try:
        accept_button = driver.find_element(by=By.XPATH, value='/html/body/div/div/div/main/form/div[1]/button[1]/span')
        time.sleep(2)
        accept_button.click()     
    except:
        pass
    try:
        accept_button = driver.find_element(by=By.XPATH, value='/html/body/div[3]/div[2]/div[1]/div[2]/div[2]/button[2]')
        time.sleep(2)
        accept_button.click()     
    except:
        pass
        


#Helper Function: get_tournament_results
# returns TRUE if all individual words are in both player and target.
def check_all_names(player,target):
    names = str(player).split()
    is_target = True
    for name in names:
        if not (name.lower() in target.lower()):
            is_target = False
    return is_target


#Helper Function: get_tournament_results
# Determines the Tournament Catagory -> U11 Gold
def find_type(match):
    out = ''
    if 'U11' in match.tournament:
        out = out + 'U11'
    if 'U13' in match.tournament:
        out = out + 'U13'
    if 'U15' in match.tournament:
        out = out + 'U15'
    if 'U17' in match.tournament:
        out = out + 'U17'
    if 'U19' in match.tournament:
        out = out + 'U19'
    if 'Senior' in str(match.tournament).lower():
        out = out + 'Senior'

    if 'Restricted' in str(match.tournament).lower():
        out = out + 'Restricted'

    if 'gold' in str(match.tournament).lower():
        out = out + ' Gold'
    if 'silver' in str(match.tournament).lower():
        out = out + ' Silver'
    if 'bronze' in str(match.tournament).lower():
        out = out + ' Bronze'
    if 'tier 4' in str(match.tournament).lower():
        out = out + 'Tier 4'
    if 'national' in str(match.tournament).lower():
        out = out + ' Nationals'
    
    return out


#Takes Driver Dicipline AgeGroup:
# Driver -> The driver being used
# Dicipline -> The dicipline wanted LOWER CASE -> ms, xd... 

#Returns RankPage(obj) with url week_values link_values
# str Url -> the url for the rank page
# date[] week_values -> the dates when the rankings have been updated
# str[] link_values -> the coresponding part of the url which changes -> 'id=[link_values[i]]&'
def initialise_links(driver,dicipline):

    class RankPage():
        def __init__(self,url,week_values,link_values,xpath):
            self.url = url
            self.week_values = week_values
            self.link_values = link_values
            self.xpath_to_div = xpath
    week_nums_id = re.compile('value="\d{4,5}">\d{1,2}-[0-9][0-9][0-9][0-9]')

    #os ws od wd mxd wxd
    ranking_links = ['https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=574&C574CS=0&C574FTYAF=0&C574FTYAT=0&C574FOG_2_F512=&p=1&ps=100',
                       'https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=575&C575CS=0&C575FTYAF=0&C575FTYAT=0&C575FOG_2_F512=&p=1&ps=100',
                       'https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=576&C576CS=0&C576FTYAF=0&C576FTYAT=0&C576FOG_2_F512=&p=1&ps=100',
                       'https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=577&C577CS=0&C577FTYAF=0&C577FTYAT=0&C577FOG_2_F512=&p=1&ps=100',
                       'https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=578&C578CS=0&C578FTYAF=0&C578FTYAT=0&C578FOG_2_F512=&p=1&ps=100',
                       'https://be.tournamentsoftware.com/ranking/category.aspx?id=37123&category=579&C579CS=0&C579FTYAF=0&C579FTYAT=0&C579FOG_2_F512=&p=1&ps=100']
                    
    xpaths = ['//*[@id="cphPage_cphPage_cphPage_C574CS_chosen"]',
              '//*[@id="cphPage_cphPage_cphPage_C575CS_chosen"]',
              '//*[@id="cphPage_cphPage_cphPage_C576CS_chosen"]',
              '//*[@id="cphPage_cphPage_cphPage_C577CS_chosen"]',
              '//*[@id="cphPage_cphPage_cphPage_C578CS_chosen"]',
              '//*[@id="cphPage_cphPage_cphPage_C579CS_chosen"]']
    if dicipline == 'os' or dicipline == 'ms':
        rank_link = ranking_links[0]
        xpath_to_div = xpaths[0]
    elif dicipline == 'ws':
        rank_link = ranking_links[1]
        xpath_to_div = xpaths[1]
    elif dicipline == 'od' or dicipline == 'md':
        rank_link = ranking_links[2]
        xpath_to_div = xpaths[2]
    elif dicipline == 'wd':
        rank_link = ranking_links[3]
        xpath_to_div = xpaths[3]
    elif dicipline == 'mxd':
        rank_link = ranking_links[4]
        xpath_to_div = xpaths[4]
    elif dicipline == 'wxd':
        rank_link = ranking_links[5]
        xpath_to_div = xpaths[5]
    else:
        return f"Can't Find Rank For Event {dicipline}"
    driver.get(rank_link)
    deal_with_cookies(driver)
    source = driver.page_source
    soup = BeautifulSoup(source,features='html.parser')

    weeks = soup.find_all('option')
    weeks_ = re.findall(week_nums_id,str(weeks))
    link_value = []
    week_date = []
    for string in weeks_:
        info = str(string).split(">")
        link_info = str(info[0]).replace('value="','').replace('"','')
        link_value.append(link_info)
        if str(info[1])[:2] == '53':
            date_in = '0-52-' + str(info[1])[-4:] 
        else:
            date_in = '0-'  + info[1]
        date_week = datetime.strptime(date_in,'%w-%W-%Y')
        week_date.append(date_week)
    return RankPage(rank_link,week_date,link_value,xpath_to_div)
    
#Takes Driver RankPage(obj) rankWeek:
# Driver -> The driver being used
# RankPage -> Obj containing rankpage url, and two arrays of weeks where the rankings updated, and their specific url ID.
# Rank Week -> Str of the date in which you would like to check the rankings at.
#Updates driver to show top 100 for the dicipline at the time specified.
def filter_ranklist(driver,RankPage,rank_week,age_group,need_ranks=True):
    url_week_regex = re.compile('id=.*?&')

    url = RankPage.url
    week_wanted = datetime.strptime(rank_week,'%d/%m/%Y')
    dates = RankPage.week_values
    ids = RankPage.link_values
    xpath = RankPage.xpath_to_div
    found = False
    for i in range(0,len(dates)):
        if dates[i] <= week_wanted and not found:
            _id = ids[i]
            found = True
    
    new_id = 'id=' +_id+ '&' 
    new_url = re.sub(url_week_regex,new_id,url)
    driver.get(new_url)
    deal_with_cookies(driver)

    if need_ranks and not (age_group == 'Senior'):
        div = driver.find_element(By.XPATH,xpath)

        div.click()
        ActionChains(driver)\
                    .move_to_element_with_offset(div,0,0)\
                    .click()\
                    .perform()


        pag.press('down', presses = how_many_presses(age_group))
        pag.press('enter')
    
        _filter = driver.find_element(By.XPATH,'//*[@id="cphPage_cphPage_cphPage_btnFilter"]')
        _filter.click()

#Helper Function: filter_agegroup
# Decides how many times the down arrow must be pressed, when selecting the Age Group Web-side
def how_many_presses(age_group):


    if age_group == 'U10':
        return 2
    if age_group == 'U11':
        return 3
    if age_group == 'U12':
        return 4
    if age_group == 'U13':
        return 5
    if age_group == 'U14':
        return 6
    if age_group == 'U15':
        return 7
    if age_group == 'U16':
        return 8
    if age_group == 'U17':
        return 9
    if age_group == 'U18':
        return 10
    if age_group == 'U19':
        return 11
    if age_group == 'Senior':
        return 12
    return 0
    

#takes overview url, player name, number of years checked.
def get_tournament_results(driver,ts_url,player_name,from_date):
    title_id = re.compile('title=.*?>')
    dates_id = re.compile('[0-9]{1,4}[\_|\-|\/|\|][0-9]{1,2}[\_|\-|\/|\|][0-9]{1,4}')
    href_id = re.compile('/player-profile/.*?/tournaments/....')
    match_date_id = re.compile('[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]')
    dicipline_id = re.compile('>Event:.*?<')
    time_id = re.compile('[0-9]m')

    tournament_url = ts_url + '/tournaments/' + str(date.today())[:4]
    driver.get(tournament_url)
    deal_with_cookies(driver)
    source = driver.page_source
    soup = BeautifulSoup(source,features='html.parser')
    years_played_html = soup.find_all('ul',id='tabs_tournaments')
    links = re.findall(href_id,str(years_played_html))
    years = []
    for link in links:
        year_in = str(link)[-4:]
        if year_in.isnumeric() and year_in not in years:
            years.append(str(link)[-4:])

    match_list = []
    tourneys = []
    tourney_dates = []
    diciplines = []

    last_known_game = int(from_date[-4:])
    for year_gathered in years:
        if int(year_gathered) > last_known_game -1:
            url = ts_url + '/tournaments/' + year_gathered
            driver.get(url)
            source = driver.page_source
            soup = BeautifulSoup(source,features='html.parser')
            tourneys_html = soup.find_all('h4')
            tourney_titles = re.findall(title_id,str(tourneys_html))
            for title in tourney_titles:
                title = title[6:-1]
                tourneys.append(title)

            tourney_dates_html = soup.find_all('div',class_='media__content')
            for dates in tourney_dates_html:
                date_ = re.findall(dates_id,str(dates))
                tourney_dates.append(date_)

            dicipline_html = soup.find_all('h4',class_='module-divider')
            for d in dicipline_html:
                dicipline = re.findall(dicipline_id,str(d))[0]
                dicipline.replace('<','')
                diciplines.append(dicipline)
            info = soup.find_all('div',class_='match')
            for input in info:
                match_info = []
                useful_info = input.text.replace('\n', ' ')
                input_data = re.split('\s',useful_info)

                k = 0
                #parse the data such its a list of useful strings
                Bye = 0
                Walkover = 0
                Cancel = 0
                while not(str(input_data[k]) == 'H2H') and not(Bye or Walkover or Cancel):
                    if str(input_data[k]) == 'Bye':
                        k = k + 1
                        Bye = 1
                    elif str(input_data[k]) == 'Walkover':
                        k = k + 1
                        Bye = 1
                    elif str(input_data[k]) == 'Statistics':
                        k = k + 1
                    elif str(input_data[k]) == 'Retired':
                        k = k + 1
                        Bye = 1
                    elif str(input_data[k]) == 'No':
                        k += 2  
                        Bye = 1                
                    elif input_data[k]:
                        name = input_data[k]
                        while input_data[k+1]:
                            name = str(name) + ' ' + str(input_data[k+1])
                            k = k + 1
                        match_info.append(name)
                        k = k + 1
                    if input_data[k] == match_date_id:
                        match_info.append(str(input_data[k]))
                    else:
                        k = k + 1
                    if(k == len(input_data) - 1):
                        Cancel = 1

                while k < len(input_data):
                    if re.findall(match_date_id,str(input_data[k])):
                        match_info.append(str(input_data[k]))
                    k += 1
                if Bye:
                    match_info = []

                if match_info:
                    match = Match()

                    #round
                    match.round = match_info[0]

                    #date
                    if re.findall(match_date_id,match_info[len(match_info)-1]):
                        match.date = match_info[len(match_info)-1]

                    #outcome and fabricate player list
                    if re.findall(time_id,match_info[1]):
                        i = 2
                    else:
                        i = 1
                    count = 0
                    player_list = []
                    while not match_info[i].isnumeric():
                        if match_info[i] == 'W':
                            match.won = True
                        elif match_info[i] == 'L':
                            match.won = False
                        else:
                            player_list.append(match_info[i])
                        count += 1
                        i += 1

                    #is_singles
                    if count == 5:
                        match.is_singles = False
                    if count == 3:
                        match.is_singles = True

                    #opponent if SINGLES
                    if match.is_singles:
                        list_of_names = []
                        opponent = ''
                        index = -1
                        for player in player_list:
                            is_player = check_all_names(player,player_name)
                            index += 1
                            if not is_player:
                                opponent = player
                            else:
                                singles_index = index
                        match.opponent = opponent

                    #opponent if DOUBLES
                    if not match.is_singles:
                        for i in range(0,4):
                            if check_all_names(player_list[i],player_name):
                                player_index = i
                        if player_index == 0:
                            match.partner = player_list[1]
                            match.opponent = player_list[2] + ' and ' + player_list[3]
                        if player_index == 1:
                            match.partner = player_list[0]
                            match.opponent = player_list[2] + ' and ' + player_list[3]
                        if player_index == 2:
                            match.partner = player_list[3]
                            match.opponent = player_list[0] + ' and ' + player_list[1]
                        if player_index == 3:
                            match.partner = player_list[2]
                            match.opponent = player_list[0] + ' and ' + player_list[1]

                    #in_three
                    score_list = []
                    for info in match_info:
                        if str(info).isnumeric():
                            score_list.append(info)
                    if len(score_list) == 6:
                        match.in_three = True

                    #correct score list if doubles
                    if not match.is_singles:
                        output = ''
                        if player_index > 1:
                            for i in range(0,len(score_list),2):
                                output = output+' '+score_list[i+1]+'-'+score_list[i]
                        else:
                            for i in range(0,len(score_list),2):
                                output = output+' '+score_list[i]+'-'+score_list[i+1]
                        match.scores = output[1:]

                    #scores if singles
                    if match.is_singles:
                        output = ''
                        if singles_index == 0:
                            for i in range(0,len(score_list),2):
                                output = output+' '+score_list[i]+'-'+score_list[i+1]
                        else:
                            for i in range(0,len(score_list),2):
                                output = output+' '+score_list[i+1]+'-'+score_list[i]
                        match.scores = output[1:]
                    match_list.append(match)

    start_dates = []
    for date_ in tourney_dates:
        if date_:
            start_dates.append(date_[1])
    
    #append tournament name to matchs

    for match in match_list:
        not_collected = True
        if match.date:
            match_date = datetime.strptime(match.date, "%d/%m/%Y")
            if match_date >= datetime.strptime(start_dates[0], "%d/%m/%Y"):
                        match.tournament = tourneys[0]
                        not_collected = False
            for i in range(1,len(start_dates)):
                if i < len(start_dates) and not_collected:
                    prev_tournament = datetime.strptime(start_dates[i], "%d/%m/%Y")
                    next_tournemant = datetime.strptime(start_dates[i-1], "%d/%m/%Y")
                    if (match_date >= prev_tournament) and match_date <= next_tournemant and not_collected:
                        match.tournament = tourneys[i]
                        not_collected = False
        else:
            match.date = 'not known'
            match.tournament = 'not known'
            

    i = 0
    first_checked = False
    from_not_known = False
    for m in match_list:
        if m.tournament == 'not known' and not first_checked:
            m.tournament = tourneys[i]
            m.date = start_dates[i]
        elif m.tournament == 'not known' and first_checked:
            m.tournament = tourneys[i+1]
            m.date = start_dates[i+1]
            from_not_known = True
        elif not m.tournament == tourneys[i]:
            if from_not_known:
                i += 1
                from_not_known = False
            i += 1
            first_checked = True
        else:
            from_not_known = False
            pass

    for match in match_list:
        type = find_type(match)
        age_group = get_agegroup(match.tournament)
        match.level = type
        match.age_group = age_group
    return match_list

#Unfinished AS FUCK
def add_diciplines_NOT_WORKING(match_list,diciplines):
    rev_matches = []
    for x in match_list:
        rev_matches.append('')
    index = len(rev_matches)-1
    for m in match_list:
        rev_matches[index] = m
        index -= 1

    def pick_dicipline(dicipline):
        event = str(dicipline).lower().replace('*','')
        if ('bs' in event) or ('os' in event) or ('ms' in event) or ('singles' in event) or ('gs' in event)  or ('ws' in event) or ('ls' in event):
            return ' Singles'
        if ('xd' in event) or ('mixed' in event):
            return 'Mixed'
        if ('bd' in event) or ('od' in event) or ('md' in event) or ('doubles' in event) or ('gd' in event)  or ('wd' in event) or ('ld' in event):
            return 'Doubles'

    def out_of_group(round):
        if 'round of' in str(round).lower():
            return True
        if 'round' in str(round).lower():
            return False
        return True
    
    def early_round(round):
        if ('Round 3' in str(round)) or ('16' in str(round)):
            return 'Round 3 Round of 16'
        if ('Round 1' in str(round)) or ('64' in str(round)):
            return 'Round 1 Round of 64'
        if ('Round 2' in str(round)) or ('32' in str(round)):
            return 'Round 2 Round of 32'
        if ('Round 4' in str(round)) or ('128' in str(round)):
            return 'Round 4 Round of 128'
        if ('Round 5' in str(round)) or ('256' in str(round)):
            return 'Round 5 Round of 256'
        

    j = len(diciplines)-1
    backtrack_to_next = False
    rounds_played = ''
    for m in rev_matches:
        if (out_of_group(m.round) and (not m.won)) or ((m.round == 'Final') and m.won) and not backtrack_to_next:
            print('here')
            m.dicipline = pick_dicipline(diciplines[j])
            backtrack_to_next = True

        elif backtrack_to_next and out_of_group(m.round):
            print('there')
            m.dicipline = pick_dicipline(diciplines[j])

        elif backtrack_to_next and not out_of_group(m.round):
            print('top of section')
            backtrack_to_next = False
            j -= 1
            m.dicipline = pick_dicipline(diciplines[j])
            rounds_played = early_round(m.round)

        
        elif 'round' in str(m.round).lower():
            print('round')
            if str(m.round).lower() in rounds_played.lower():
                j -= 1
                print(m.scores,m.date)
                m.dicipline = pick_dicipline(diciplines[j])
                rounds_played = ''
            else:
                rounds_played = rounds_played + ' ' + early_round(m.round)
                print('just a round')
                m.dicipline = pick_dicipline(diciplines[j])
        print(m.round, ' == ', rounds_played)

    
    for match in match_list:
        print(match.scores,match.dicipline)

    ##UNFINISHES

def add_diciplines(matches):
    for m in matches:
        if m.is_singles:
            m.dicipline = 'Singles'
        elif m.is_level_dubs:
            m.dicipline = 'Doubles'
        else:
            m.dicipline = 'Mixed'

def load_senior_ranks():
    diciplines = ['ms','ws','md','wd','mxd','wxd']
    rank_files = []
    for dicipline in diciplines:
        path = 'Ranks/'+dicipline+'.xlsx'
        xl = pd.ExcelFile(path)
        rank_files.append(xl)
    return rank_files

def parse_ranks_from_excel(rank_files,dicipline,date):
    if dicipline == 'ms': xl = rank_files[0]
    elif dicipline == 'ws': xl = rank_files[1]
    elif dicipline == 'md': xl = rank_files[2]
    elif dicipline == 'wd': xl = rank_files[3]
    elif dicipline == 'mxd': xl = rank_files[4]
    elif dicipline == 'wxd': xl = rank_files[5]
    else: return 'Error'

    rank_sheets = xl.sheet_names
    rank_sheets.reverse()
    for i in range(0,len(rank_sheets)):
        week_date = '1-'+rank_sheets[i]
        week_num = datetime.strptime(week_date,'%w-%W-%Y')
        if week_num < date:
            df = xl.parse(rank_sheets[i])
            return df['players'].tolist()
    return 'Error'

def add_ranks_to_matches(driver,matches,player_name,gender,age_group):

    rank_sheets = load_senior_ranks()

    def filter_and_get_ranks(driver,dicipline,rank_page,date,age_group):

        if not age_group=='Senior':
            date = datetime.strftime(date,'%d/%m/%Y')
            filter_ranklist(driver,rank_page,date,age_group)
            return get_ranks(driver,rank_page,date,age_group)
        else:
            return parse_ranks_from_excel(rank_sheets,dicipline,date)

    if gender == 'm':
        links_to_check = ['ms','md','mxd','wxd']
    if gender == 'w':
        links_to_check = ['ws','wd','mxd','wxd']

    today = datetime.today()
    rank_page_singles = initialise_links(driver,links_to_check[0])
    rank_page_dubs = initialise_links(driver,links_to_check[1])
    rank_page_mxd = initialise_links(driver,links_to_check[2])
    rank_page_wxd = initialise_links(driver,links_to_check[3])


    check_ranks_mxd = parse_ranks_from_excel(rank_sheets,'mxd',today)
    check_ranks_wxd = parse_ranks_from_excel(rank_sheets,'wxd',today)
    
    clumps = group_matches_by_date(driver,matches,rank_page_singles)
    for clump in clumps:
        clump_date = clump.week_value    #datetime.strftime(clump.week_value,'%d/%m/%Y')
        check_first_in_clump = False
        for i in range(0,len(clump.games)):
            match = clump.games[i]
            match.ranks_used = age_group
            if not check_first_in_clump:
                    same_time = False
                    check_first_in_clump = True
                    same_age = False
            else:
                if (match.tournament == clump.games[i-1].tournament) or match.tournament == 'not known':
                    same_time = True 
                    same_age = True
                else:
                    if match.age_group == clump.games[i-1].age_group:
                        same_age = True
                    else:
                        same_age = False
                    same_time = False
                #if same time -> DONT re check the ranks. 
                #if not the same -> reset ranks if age group changes
                #hard reset for next tournament
            if not same_time and not same_age:
                ranks_mxd = filter_and_get_ranks(driver,links_to_check[2],rank_page_mxd,clump_date,age_group)
                ranks_wxd = filter_and_get_ranks(driver,links_to_check[3],rank_page_wxd,clump_date,age_group)
                ranks_dubs = filter_and_get_ranks(driver,links_to_check[1],rank_page_dubs,clump_date,age_group)
                ranks_singles = filter_and_get_ranks(driver,links_to_check[0],rank_page_singles,clump_date,age_group)
            if match.is_singles:
                match.opponent_rank = find_rank(match.opponent,ranks_singles,match.is_singles)
                match.player_rank = find_rank(player_name,ranks_singles,match.is_singles)
            else:
                #Player Is A Bloke
                if links_to_check[0] == 'ms':
                    partner_rank = find_rank(match.partner,check_ranks_wxd,match.is_singles)
                    if not partner_rank:
                    #Assume that its probably Mens Dubs
                        match.partner_rank = find_rank(match.partner,ranks_dubs,match.is_singles)
                        match.is_level_dubs = True
                    else:
                        match.partner_rank = partner_rank
                        match.is_level_dubs = False

                    if match.is_level_dubs:
                        match.opponent_rank = find_rank(match.opponent,ranks_dubs,match.is_singles)
                        match.player_rank = find_rank(player_name,ranks_dubs,match.is_singles)
                    else:
                        match.woman_rank = find_rank(match.opponent,ranks_wxd,match.is_singles)
                        match.man_rank = find_rank(match.opponent,ranks_mxd,match.is_singles)
                        match.opponent_rank = [str(match.man_rank),str(match.woman_rank)] 
                        match.player_rank = find_rank(player_name,ranks_mxd,match.is_singles)

                #Player Is A Shelia
                if links_to_check[0] == 'ws':
                    partner_rank = find_rank(match.partner,check_ranks_mxd,match.is_singles)
                    if not partner_rank:
                    #Assume that its probably Womens Dubs
                        match.partner_rank = find_rank(match.partner,ranks_dubs,match.is_singles)
                        match.is_level_dubs = is_level_dubs = True
                    else:
                        match.partner_rank = find_rank(match.partner,ranks_mxd,match.is_singles)
                        match.is_level_dubs = is_level_dubs = False
                    if match.is_level_dubs:
                        match.opponent_rank = find_rank(match.opponent,ranks_dubs,match.is_singles)
                        match.player_rank = find_rank(player_name,ranks_dubs,match.is_singles)
                    else:
                        match.woman_rank = find_rank(match.opponent,ranks_wxd,match.is_singles)
                        match.man_rank = find_rank(match.opponent,ranks_mxd,match.is_singles)
                        match.opponent_rank = [str(match.woman_rank),str(match.man_rank)] 
                        match.player_rank = find_rank(player_name,ranks_wxd,match.is_singles)
    add_diciplines(matches)
    return matches

def add_dicipline_not_ranks_to_matches(driver,matches,ts_url,player_name,gender,age_group):

    rank_page_mxd = initialise_links(driver,'mxd')
    rank_page_wxd = initialise_links(driver,'wxd')
    
    today = datetime.strftime(datetime.today(),'%d/%m/%Y')
    
    filter_ranklist(driver,rank_page_mxd,today,'Senior',need_ranks=False)
    check_ranks_mxd = get_ranks(driver,rank_page_mxd,today,'Senior')
    filter_ranklist(driver,rank_page_wxd,today,'Senior',need_ranks=False)
    check_ranks_wxd = get_ranks(driver,rank_page_wxd,today,'Senior')
    
    for match in matches:
        if not match.is_singles:
            if gender == 'm':
                #Player is a bloke
                    partner_rank = find_rank(match.partner,check_ranks_wxd,match.is_singles)
                    if not partner_rank:
                    #Assume that its probably Mens Dubs
                        match.is_level_dubs = True
                    else:
                        match.is_level_dubs = False

                #Player Is A Shelia
            if gender == 'w':
                partner_rank = find_rank(match.partner,check_ranks_mxd,match.is_singles)
                if not partner_rank:
                #Assume that its probably Womens Dubs
                    match.is_level_dubs = is_level_dubs = True
                else:
                    match.is_level_dubs = is_level_dubs = False
    
    add_diciplines(matches)
    return(matches)

def get_ranks(driver,RankPage,rank_week,age_group):

    page_nums = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30]

    name_regex = re.compile('href=.*?</a>')
    next_page_regex = re.compile('href=.*?title')
    deal_with_cookies(driver)
    
    useful_names = []
    for page in page_nums:
        try:
            if not(page == 1):
                source = driver.page_source
                soup = BeautifulSoup(source,features='html.parser')
                next_button = soup.find_all('a',class_='page_next')
                next = re.findall(next_page_regex,str(next_button))
                next = str(next[0]).replace('href="','https://be.tournamentsoftware.com').replace('" title','').replace('amp;','').replace('ps=25','ps=100')
                driver.get(next)
            source = driver.page_source
            soup = BeautifulSoup(source,features='html.parser')
            names_html = soup.find_all('table',class_='ruler')
            names = re.findall(name_regex,str(names_html)) 
            if names:
                for i in range(6,len(names)):
                    name = str(names[i])
                    if not('Profile' in name) and not('category' in name):
                        name = name.split('>')[1].replace('</a','')
                        useful_names.append(name)
        except:
            pass
    return useful_names

def find_rank(opponent,ranks,is_singles):
    if 'and' in opponent:
        two_names = True
    else:
        two_names = False
    i = 1
    not_found = ''

    if is_singles:
        for player in ranks:
            if check_all_names(player,opponent):
                return i
            i+=1
        return not_found
    
    else:
        if two_names:
            opponents = str(opponent).split(' and ')
        out = []
        both_found = 0
        if two_names:
            for player_ in opponents:
                notFound = True
                for player in ranks:
                    if check_all_names(player_,player) and notFound:
                        out.append(str(i))
                        both_found += 1
                        notFound = False
                    i+=1
                i = 1
        else:
            for player in ranks:
                    if check_all_names(opponent,player):
                        out.append(str(i))
                        both_found += 1
                    i+=1

        if both_found == 2 and two_names:
            return out
        elif both_found == 1 and not two_names:
            return out
        elif both_found == 1 and two_names:
            return out


#player_name = 'Sajan Senthuran'
#ts_url = 'https://be.tournamentsoftware.com/player-profile/9258742A-9538-4CA7-842B-02B56A3D6590'
#age_group = 'Senior'
#driver = webdriver.Firefox()
#initialise_links(driver,'os',age_group)
#get_tournament_results(driver,ts_url,player_name)
#driver.close()

#Helper Function: group_matches_by_date
#Returns true if the date is between lower(older) and upper(more recent). replace with 0 for one sided check.
#IMPUT Datetime opjects, not strings.
def date_in_range(date,lower,upper):
    if upper == 0:
        upper = datetime.today()
    if lower == 0:
        lower = datetime.strptime('1980','%Y') 
    if date <= upper and date >= lower:
        return True
    return False


#Takes Driver Matches(obj)[] RankPage(obj):
# Driver -> The driver being used
# Matches -> Obj containing a list of Match opjects
# Rank Data -> RankWeek Obj from rank_page()

#Returns MatchClump(Obj)[] with: games week_value link
# Games -> the games played in that 'Clump'
# week_value -> the most recent rank update date
# link -> the link to that week in the ranking page
def group_matches_by_date(driver,matches,rank_data):
    all_games = []
    matchs_in_period = []
    first = True
    first_period_found = False
    i = 0
    week_values,link_values = rank_data.week_values,rank_data.link_values

    for match in matches:
        if str(match.date) == 'not known':
            matchs_in_period.append(match)
            has_date = False
        else:
            match_date = datetime.strptime(match.date,'%d/%m/%Y')
            has_date = True
        
        if has_date:
            if first:
                matchs_in_period.append(match)        
                first = False
            else:
                if date_in_range(match_date,week_values[i],week_values[i-1]):
                    matchs_in_period.append(match)
                    first_period_found = True
                else:
                    if matchs_in_period and first_period_found:
                        all_games.append(MatchClump(matchs_in_period,week_values[i],link_values[i]))
                        matchs_in_period = []
                        matchs_in_period.append(match)
                    else:
                        matchs_in_period.append(match)
                    i += 1
    if matchs_in_period:            
        all_games.append(MatchClump(matchs_in_period,week_values[i],link_values[i]))     

            
    return all_games

def get_agegroup(title):
    age_regex = re.compile('U1[0-9]')
    age = re.findall(age_regex,title)
    if age:
        return str(age[0])
    elif 'Senior' in title:
        return 'Senior'
    return 'Unknown age'

def add_extras(matches,age_group):
    tournament_info = tournament_runs(matches,age_group)

def tournament_runs(matches,age_group,dicipline):
    results = [0,0,0,0,0,0]
    wins = []
    new_tourney = 0
    for m in matches:
        
        if dicipline == 'singles':
            check = m.is_singles
        if dicipline == 'dubs':
            check = m.is_level_dubs
        if dicipline == 'mixed':
            check = not m.is_level_dubs

        if str(m.age_group) == age_group and check:

            if new_tourney == 0:
                new_tourney = str(m.tournament)
            elif new_tourney != str(m.tournament):
                results[0] += 1
                new_tourney = str(m.tournament)

            if '16' in str(m.round):
                results[1] += 1
            if 'Quarter' in str(m.round):
                results[2] += 1
            if 'Semi' in str(m.round):
                results[3] += 1
            if 'Final' in str(m.round):

                results[4] += 1
                if m.won:
                    results[5] += 1
                    wins.append(m.tournament + '- Winner')
                else:
                    wins.append(m.tournament + ' - Runner Up')
    return results,wins

def wins_vs_rank(matches,age_group):
    singles = [0,0]
    doubles = [0,0]
    mixed = [0,0]
    for m in matches:
        if m.age_group == age_group:
            if m.is_singles:
                if m.won:
                    if m.player_rank > m.opponent_rank:
                        singles[0] += 1

def make_df(matches,date,with_ranks):
    _tournament = []
    _level = []
    _dicipline = []
    _round = []
    _outcome = []
    _partner = []
    _home_ranks = []
    _opponent = []
    _away_ranks = []
    _score = []
    _date = []

    colums = ['Date','Tournament','Dicipline','Round','Scores','Opponent/s','Outcome','3-Sets','Partner','Level','Age','Home Ranks','Opp Ranks','Ranks Used','Year']

    dataframe = []

    date = datetime.strptime(date,'%d/%m/%Y') + timedelta(hours=48)
    for m in matches:
        no_date = False
        if m.date == 'not known':
            match_date = datetime.strptime('1/1/1975','%d/%m/%Y')
            no_date = True
        else:
            match_date = datetime.strptime(m.date,'%d/%m/%Y')

        if (match_date > date) or no_date:
            row = []
            #Date Tourney Dicipline Round Scores opponents Outcome In-3 Partner level age home-r opp-r Ranks_used Year
            row.append(m.date)
            row.append(m.tournament)
            row.append(m.dicipline)
            row.append(m.round)
            row.append(m.scores)
            row.append(m.opponent)
            
            if m.won:
                row.append('Won')
            else:
                row.append('Lost')

            if m.in_three:
                row.append('Yes')
            else:
                row.append('No')

            if m.partner:
                row.append(m.partner)
                if m.partner_rank:
                    parank = m.partner_rank[0]
                else:
                    parank = '?'
                if m.player_rank:
                    plrank = m.player_rank[0]
                else:
                    plrank = '?'
                ranks = str(plrank) + ' and ' + str(parank)
            else:   
                row.append('')
            
            row.append(m.level)
            row.append(m.age_group)
            if with_ranks:
                if m.partner:
                    row.append(ranks)
                else:
                    row.append(m.player_rank)
                print(m.opponent)
                if 'and' in  str(m.opponent):
                    try:
                        out = str(m.opponent_rank[0])+', '+str(m.opponent_rank[1])
                        row.append(str(out))
                    except:
                        if m.opponent_rank:
                            out = str(m.opponent_rank) + ', ?'
                            row.append(str(out))
                        else:
                            row.append(('?, ?'))
                else:
                    row.append(m.opponent_rank)

                row.append(m.ranks_used)
            else:
                row.append('')
                row.append('')
                row.append('')
            
            if m.date == 'not known':
                row.append('')
            else:
                row.append(str(m.date)[-4:])

            dataframe.append(row)
        
        else:
            df = pd.DataFrame(dataframe,columns=colums)
            df.fillna('')
            return df

    df = pd.DataFrame(dataframe,columns=colums)
    df.fillna('')
    return df

def tournament_totals(matches,dicipline):
    class TypeOfTournament():
        def __init__(self,name):
            self.name = name
            self.wins = []
            self.results = [0,0,0,0,0]

    t4 = TypeOfTournament('Tier 4')
    bronze = TypeOfTournament('Bronze')
    silver = TypeOfTournament('Silver')
    gold = TypeOfTournament('Gold')
    nationals = TypeOfTournament('Nationals')
    tts = [t4,bronze,silver,gold,nationals]
    last_tournament = ''
    for index,m in matches.iterrows():
        if  m['Dicipline'] == dicipline:
            should_change = check_result(m,tts,last_tournament)
            if should_change:
                last_tournament = m['Tournament']

    dfs = []
    for t in tts:
        df = make_runs_to_df(t)
        df.fillna('')
        dfs.append(df)
    return dfs[0],dfs[1],dfs[2],dfs[3],dfs[4]

def check_result(match,tts,last_tournament):
    if 'tier 4' in str(match['Level']).lower():
        tourney = tts[0]
    elif 'bronze' in str(match['Level']).lower():
        tourney = tts[1]
    elif 'silver' in str(match['Level']).lower():
        tourney = tts[2]
    elif 'gold' in str(match['Level']).lower():
        tourney = tts[3]
    elif 'national' in str(match['Level']).lower():
        tourney = tts[4]
    
    else:
        tourney = ''

    if str(match['Tournament']) != str(last_tournament):
        new_tourney = True
    else:
        new_tourney = False 

    if tourney:
        if new_tourney:
            tourney.results[0] += 1
        if 'Quarter' in str(match.round):
            tourney.results[1] += 1
        if 'Semi' in str(match.round):
            tourney.results[2] += 1
        if 'Final' in str(match.round):
            tourney.results[3] += 1
            if match['Outcome'] == 'Won':
                tourney.results[4] += 1
                tourney.wins.append(str(match['Tournament']) + '- Winner - ' + str(match['Date'])[-4:])
            else:
                tourney.wins.append(str(match['Tournament']) + ' - Runner Up - ' + str(match['Date'])[-4:])
        return True
    return False

def get_ko_results(matches):
    class KoMatch():
        def __init__(self,round,opponent,tournament,age_group,date,dicipline,outcome):
            self.round = round
            self.opponent = opponent
            self.dicipline = dicipline
            self.partner = ''
            self.tournament = tournament
            self.date = date
            self.outcome = outcome
            self.age_group = age_group
    ko_matches = []
    for m in matches:
        if not 'Round' in str(m.round):
            out = str(m.round).split()
            round_played = out[0]

            if m.won:
                outcome = 'Won'
            else:
                outcome = 'Lost'
            
            if m.is_singles:
                dicipline = 'Singles'
                need_partner = False
            elif m.is_level_dubs:
                dicipline = 'Doubles'
                need_partner = True
            else:
                dicipline = 'Mixed'
                need_partner = True

            if need_partner:
                partner = m.partner

            age_group = m.age_group
            opponent = m.opponent
            tournament = m.tournament
            date = m.date

            match = KoMatch(round_played,opponent,tournament,age_group,date,dicipline,outcome)
            if need_partner:
                match.partner = partner
            ko_matches.append(match)

    return ko_matches

def make_runs_to_df(tourney_level):
    t = tourney_level 
    input = t.results
    wins = ''
    if t.wins:
        for w in t.wins:
            wins = wins + ' | ' + w
        input.append(wins)
    else:
        input.append('None')

    df = pd.DataFrame([input],columns=['Played','Quaters','Semis','Finals','Won','Notable Results'])
    df.style.set_caption(t.name)

    return df

def update_ranks(dicipline):
    #backup old ranks incase of corruption
    os.remove('Ranks/Backup')
    os.makedirs('Ranks/Backup')
    shutil.copy('Ranks/md.xlsx','Ranks/Backup/md.xlsx')
    shutil.copy('Ranks/wd.xlsx','Ranks/Backup/wd.xlsx')
    shutil.copy('Ranks/ws.xlsx','Ranks/Backup/ws.xlsx')
    shutil.copy('Ranks/ms.xlsx','Ranks/Backup/ms.xlsx')
    shutil.copy('Ranks/mxd.xlsx','Ranks/Backup/mxd.xlsx')
    shutil.copy('Ranks/wxd.xlsx','Ranks/Backup/wxd.xlsx')

    driver = webdriver.Firefox()
    RankPage = initialise_links(driver,dicipline)
    path = 'Ranks/'+dicipline+'.xlsx'

    book = load_workbook(path)

    last_update = '1-' + book.sheetnames[0]

    earliest_date = datetime.strptime(last_update,'%w-%U-%Y') #%d/%m/%Y %U-%Y
  

    
    #convert sheet names to good datetime - DECENDING - 2024, 2023... 
    dates = []
    for date in reversed(book.sheetnames):
        dtDate = datetime.strptime('1-' + date,'%w-%U-%Y')
        dates.append(dtDate)

    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    for week in reversed(RankPage.week_values):
        sheet_name = datetime.strftime(week,'%w-%W-%Y')
        sheet_name = sheet_name[2:]
        if (not (sheet_name in book.sheetnames)) and week > earliest_date:
            week_date = datetime.strftime(week,'%d/%m/%Y')
            filter_ranklist(driver,RankPage,week_date,'Senior')
            ranks = get_ranks(driver,1,1,1)
            df = pd.DataFrame(ranks,columns=['players'])
            df.to_excel(writer,sheet_name=sheet_name)

    writer.close()         
    driver.close()   

def generate_spreadsheet(player_name,ts_url,age_group,gender,sheet,with_ranks=False):
    sheet = 'Players/' + sheet

    driver = webdriver.Firefox()
    matches = get_tournament_results(driver,ts_url,player_name,'1/1/2015')
    if with_ranks:
        matches = add_ranks_to_matches(driver,matches,player_name,gender,age_group)
    else:
        matches = add_dicipline_not_ranks_to_matches(driver,matches,ts_url,player_name,gender,age_group)
    driver.close()
    if not exists(sheet):
        shutil.copy('Util/Template.xlsx',sheet)
        frame = make_df(matches,'1/06/2015',with_ranks)
        frame.to_excel('Util/temp.xlsx',index=0)
        matches_ = pd.read_excel('Util/temp.xlsx',usecols='A:O',keep_default_na=False)
    else:
        old_matches = pd.read_excel(sheet,sheet_name='MatchInput',usecols='A:O',keep_default_na=False)
        new_matches = make_df(matches,str(old_matches['Date'][1]),with_ranks)
        if not new_matches.empty:
            frame = pd.concat([new_matches,old_matches],ignore_index=True)
        else:
            frame = old_matches
        frame.fillna('')
        matches_ = frame

    s_tier4_df,s_bronze_df,s_silver_df,s_gold_df,s_nationals_df = tournament_totals(matches_,'Singles')
    d_tier4_df,d_bronze_df,d_silver_df,d_gold_df,d_nationals_df = tournament_totals(matches_,'Doubles')
    x_tier4_df,x_bronze_df,x_silver_df,x_gold_df,x_nationals_df = tournament_totals(matches_,'Mixed')

    dfs = [s_tier4_df,s_bronze_df,s_silver_df,s_gold_df,s_nationals_df,d_tier4_df,d_bronze_df,d_silver_df,d_gold_df,d_nationals_df,x_tier4_df,x_bronze_df,x_silver_df,x_gold_df,x_nationals_df]

    workbook = openpyxl.load_workbook(sheet)
    wsheet = workbook['MatchInput']
    wsheet.delete_rows(1,1000)
    for row in dataframe_to_rows(matches_,index=False):
        wsheet.append(row)

    wsheet = workbook['ExtraInput']
    wsheet.delete_rows(1,300)
    for df in dfs:
        for row in dataframe_to_rows(df,index=False):
            wsheet.append(row)
    workbook.save(sheet)
    workbook.close()
#----------------------------------------------------------------#



#----------------------------------------------------------------#